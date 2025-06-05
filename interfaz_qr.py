import tkinter as tk
from tkinter import messagebox, filedialog, Toplevel, Listbox
import threading
from qr_scanner import iniciar_escaneo_qr, descargar_excel_desde_qr_scanner
from generar_qrs import generar_qr_para_usuario
from PIL import Image, ImageTk
import os
import shutil
from openpyxl import load_workbook
from openpyxl import Workbook
import pymysql
import zipfile
import psutil
import msvcrt

# Ruta del Excel
excel_file = "asistencias.xlsx"
qr_folder = "imagenes_qr"

# Crear la carpeta de imágenes QR
if not os.path.exists(qr_folder):
    os.makedirs(qr_folder)

def archivo_excel_abierto():
    try:
        with open(excel_file, 'r+b') as f:
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
            msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
        return False
    except (OSError, PermissionError):
        return True

# Función para iniciar la cámara y escanear QR
def iniciar_camara():
    if archivo_excel_abierto():
        messagebox.showerror("Error", "No puedes registrar asistencias mientras el archivo de Excel esté abierto. Por favor, cierra el archivo e inténtalo de nuevo.")
        return
    def ejecutar_escaneo():
        try:
            iniciar_escaneo_qr()
        except Exception as e:
            print(f"Error al iniciar la cámara: {e}")
            messagebox.showerror("Error", f"Error al iniciar la cámara: {e}")

    hilo = threading.Thread(target=ejecutar_escaneo)
    hilo.daemon = True
    hilo.start()

# Consolidar funciones repetitivas y mejorar la organización

def descargar_archivo(tipo, ruta_destino):
    try:
        if tipo == "excel":
            shutil.copy(excel_file, ruta_destino)
        elif tipo == "bd":
            with open(ruta_destino, "w") as respaldo:
                conexion = pymysql.connect(host="localhost", user="root", password="1234", database="asistencia")
                cursor = conexion.cursor()
                cursor.execute("SHOW CREATE TABLE usuarios;")
                respaldo.write(cursor.fetchone()[1] + ";\n")
                cursor.execute("SELECT * FROM usuarios;")
                for row in cursor.fetchall():
                    respaldo.write(f"INSERT INTO usuarios VALUES {row};\n")
                cursor.execute("SHOW CREATE TABLE asistencia;")
                respaldo.write(cursor.fetchone()[1] + ";\n")
                cursor.execute("SELECT * FROM asistencia;")
                for row in cursor.fetchall():
                    respaldo.write(f"INSERT INTO asistencia VALUES {row};\n")
                cursor.close()
                conexion.close()
        elif tipo == "qr":
            with zipfile.ZipFile(ruta_destino, "w") as zipf:
                for archivo in os.listdir(qr_folder):
                    ruta_archivo = os.path.join(qr_folder, archivo)
                    if os.path.isfile(ruta_archivo):
                        zipf.write(ruta_archivo, os.path.join("imagenes_qr", archivo))
        messagebox.showinfo("Éxito", f"Archivo {tipo} guardado en: {ruta_destino}")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el archivo {tipo}: {e}")

# Reemplazar funciones específicas con la función consolidada
def descargar_excel():
    ruta_destino = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")], title="Guardar archivo de Excel")
    if ruta_destino:
        descargar_archivo("excel", ruta_destino)

def descargar_bd():
    ruta_destino = filedialog.asksaveasfilename(defaultextension=".sql", filetypes=[("Archivos SQL", "*.sql")], title="Guardar respaldo de la base de datos")
    if ruta_destino:
        descargar_archivo("bd", ruta_destino)

def descargar_imagenes_qr():
    ruta_destino = filedialog.asksaveasfilename(defaultextension=".zip", filetypes=[("Archivo ZIP", "*.zip")], title="Guardar imágenes de QR como ZIP")
    if ruta_destino:
        descargar_archivo("qr", ruta_destino)

# Función para generar un nuevo usuario
def generar_usuario_desde_interfaz():
    def ejecutar_generacion():
        try:
            nombres = entry_nombre.get()
            apellido = entry_apellido.get()
            segundo_apellido = entry_segundo_apellido.get()

            if not nombres or not apellido:
                messagebox.showwarning("Advertencia", "Por favor, ingresa el nombre(s) y apellido(s).")
                return

            # Llama a la función con los tres campos
            resultado = generar_qr_para_usuario(nombres, apellido, segundo_apellido)
            messagebox.showinfo("Éxito", resultado)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el usuario: {e}")

    hilo = threading.Thread(target=ejecutar_generacion)
    hilo.daemon = True
    hilo.start()

# Función para mostrar todos los QRs existentes
def mostrar_todos_qrs():
    if not os.path.exists(qr_folder):
        messagebox.showerror("Error", f"La carpeta '{qr_folder}' no existe. Por favor, genera un QR primero.")
        return

    archivos_qr = [f for f in os.listdir(qr_folder) if f.endswith(".png")]

    if not archivos_qr:
        messagebox.showinfo("Información", "No hay códigos QR disponibles. Por favor, genera un QR primero.")
        return

    ventana_qrs = Toplevel(root)
    ventana_qrs.title("Códigos QR Disponibles")
    ventana_qrs.geometry("400x500")

    # Add a search bar
    def filtrar_qrs():
        query = entry_buscar.get().lower()
        lista_qrs.delete(0, tk.END)
        for archivo in archivos_qr:
            if query in archivo.lower():
                lista_qrs.insert(tk.END, archivo)

    label_buscar = tk.Label(ventana_qrs, text="Buscar QR:", font=("Arial", 12))
    label_buscar.pack(pady=5)
    entry_buscar = tk.Entry(ventana_qrs, width=30)
    entry_buscar.pack(pady=5)
    entry_buscar.bind("<KeyRelease>", lambda event: filtrar_qrs())

    lista_qrs = Listbox(ventana_qrs, width=50, height=20)
    lista_qrs.pack(pady=10)

    for archivo in archivos_qr:
        lista_qrs.insert(tk.END, archivo)

    def abrir_qr():
        seleccion = lista_qrs.curselection()
        if seleccion:
            archivo_seleccionado = lista_qrs.get(seleccion[0])  # Obtener el nombre directamente del Listbox
            ruta_archivo = os.path.join(qr_folder, archivo_seleccionado)
            try:
                ventana_imagen = Toplevel(ventana_qrs)
                ventana_imagen.title(archivo_seleccionado)

                imagen = Image.open(ruta_archivo)
                imagen_tk = ImageTk.PhotoImage(imagen)

                label_imagen = tk.Label(ventana_imagen, image=imagen_tk)
                label_imagen.image = imagen_tk
                label_imagen.pack()

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
        else:
            messagebox.showwarning("Advertencia", "Por favor, selecciona un QR para abrir.")

    btn_abrir_qr = tk.Button(ventana_qrs, text="Abrir QR", command=abrir_qr, bg="purple", fg="white")
    btn_abrir_qr.pack(pady=10)

# Estructura de la interfaz
root = tk.Tk()
root.title("Sistema de Asistencia con QR")
root.geometry("500x700")  # Adjusted window size

# Improve the interface design
root.configure(bg="lightblue")

label_titulo = tk.Label(root, text="Sistema de Asistencia con QR", font=("Arial", 20, "bold"), bg="lightblue", fg="darkblue")
label_titulo.pack(pady=10)

# Reorganize the interface layout
frame_botones = tk.Frame(root, bg="lightblue")
frame_botones.pack(pady=40)  # Aumentar el espaciado vertical general

btn_iniciar_camara = tk.Button(frame_botones, text="Iniciar Cámara", command=iniciar_camara, width=20, height=2, bg="green", fg="white", font=("Arial", 12, "bold"))
btn_iniciar_camara.grid(row=0, column=0, padx=10, pady=10)

# Función para limpiar los datos (versión funcional previa)
def limpiar_datos():
    ventana_opciones = Toplevel(root)
    ventana_opciones.title("Opciones de Limpieza de Datos")
    ventana_opciones.geometry("400x300")

    label = tk.Label(ventana_opciones, text="¿Qué deseas limpiar?", font=("Arial", 14))
    label.pack(pady=10)

    def advertencia_accion(tipo):
        ventana_advertencia = Toplevel(ventana_opciones)
        ventana_advertencia.title("ADVERTENCIA")
        ventana_advertencia.geometry("420x220")
        if tipo == "excel":
            desc = "Esto eliminará todas las asistencias del archivo Excel, pero conservará los usuarios y sus QRs. ¿Deseas continuar?"
        elif tipo == "asistencias":
            desc = "Esto eliminará todas las asistencias de la base de datos, pero conservará los usuarios y sus QRs. ¿Deseas continuar?"
        else:
            desc = "Esto eliminará TODOS los datos: usuarios, asistencias, Excel e imágenes QR. ¿Deseas continuar? Se recomienda descargar un respaldo."
        label_adv = tk.Label(ventana_advertencia, text=desc, wraplength=400, fg="red", font=("Arial", 11, "bold"))
        label_adv.pack(pady=10)

        def descargar():
            descargar_todo()

        def aceptar():
            if tipo == "excel":
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["Nombre", "Apellido", "Número de Alumno"])
                    wb.save(excel_file)
                    messagebox.showinfo("Éxito", "Asistencias en Excel limpiadas.")
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo limpiar el Excel: {e}")
            elif tipo == "asistencias":
                try:
                    conexion = pymysql.connect(host="localhost", user="root", password="1234", database="asistencia")
                    cursor = conexion.cursor()
                    cursor.execute("DELETE FROM asistencia;")
                    conexion.commit()
                    cursor.close()
                    conexion.close()
                    messagebox.showinfo("Éxito", "Asistencias en la base de datos limpiadas.")
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo limpiar la base de datos: {e}")
            else:
                try:
                    # Limpiar Excel
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["Nombre", "Apellido", "Número de Alumno"])
                    wb.save(excel_file)
                    # Limpiar BD
                    conexion = pymysql.connect(host="localhost", user="root", password="1234", database="asistencia")
                    cursor = conexion.cursor()
                    cursor.execute("DELETE FROM asistencia;")
                    cursor.execute("DELETE FROM usuarios;")
                    conexion.commit()
                    cursor.close()
                    conexion.close()
                    # Limpiar imágenes QR
                    for archivo in os.listdir(qr_folder):
                        ruta_archivo = os.path.join(qr_folder, archivo)
                        if os.path.isfile(ruta_archivo):
                            os.remove(ruta_archivo)
                    messagebox.showinfo("Éxito", "Todos los datos han sido limpiados.")
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo limpiar todo: {e}")
            ventana_advertencia.destroy()
            ventana_opciones.destroy()

        btn_cancelar = tk.Button(ventana_advertencia, text="Cancelar", command=ventana_advertencia.destroy, bg="gray", fg="white", width=12)
        btn_cancelar.pack(side=tk.LEFT, padx=15, pady=20)
        btn_descargar = tk.Button(ventana_advertencia, text="Descargar datos", command=descargar, bg="blue", fg="white", width=15)
        btn_descargar.pack(side=tk.LEFT, padx=15, pady=20)
        btn_aceptar = tk.Button(ventana_advertencia, text="Aceptar", command=aceptar, bg="red", fg="white", width=12)
        btn_aceptar.pack(side=tk.LEFT, padx=15, pady=20)

    btn_excel = tk.Button(ventana_opciones, text="Limpiar solo Excel (asistencias)", command=lambda: advertencia_accion("excel"), width=30, bg="orange", fg="white")
    btn_excel.pack(pady=10)
    btn_asistencias = tk.Button(ventana_opciones, text="Limpiar solo asistencias (BD)", command=lambda: advertencia_accion("asistencias"), width=30, bg="purple", fg="white")
    btn_asistencias.pack(pady=10)
    btn_todo = tk.Button(ventana_opciones, text="Limpiar TODO (Excel, BD, QRs)", command=lambda: advertencia_accion("todo"), width=30, bg="red", fg="white")
    btn_todo.pack(pady=10)
    btn_cerrar = tk.Button(ventana_opciones, text="Cerrar", command=ventana_opciones.destroy, width=30, bg="gray", fg="white")
    btn_cerrar.pack(pady=10)

# Botón para limpiar datos (posición correcta y visible)
btn_limpiar_datos = tk.Button(frame_botones, text="Limpiar Datos", command=limpiar_datos, width=20, height=2, bg="red", fg="white", font=("Arial", 12, "bold"))
btn_limpiar_datos.grid(row=0, column=1, padx=10, pady=10)

btn_descargar_excel = tk.Button(frame_botones, text="Descargar Excel", command=descargar_excel, width=20, height=2, bg="blue", fg="white", font=("Arial", 12, "bold"))
btn_descargar_excel.grid(row=1, column=1, padx=10, pady=10)  # Move "Descargar Excel" to the second row, second column

btn_visualizar_excel = tk.Button(frame_botones, text="Visualizar Excel", command=lambda: os.startfile(excel_file), width=20, height=2, bg="gray", fg="white", font=("Arial", 12, "bold"))
btn_visualizar_excel.grid(row=1, column=0, padx=10, pady=10)

# Replace the three download buttons with a single "Descargar Todo" button
def descargar_todo():
    ventana_descargar = Toplevel(root)
    ventana_descargar.title("Descargar Datos")
    ventana_descargar.geometry("400x300")
    label_titulo = tk.Label(ventana_descargar, text="Selecciona qué deseas descargar:", font=("Arial", 14))
    label_titulo.pack(pady=10)
    def descargar_excel():
        ruta_destino = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")], title="Guardar archivo de Excel")
        if ruta_destino:
            descargar_archivo("excel", ruta_destino)
    def descargar_bd():
        ruta_respaldo_db = filedialog.asksaveasfilename(defaultextension=".sql", filetypes=[("Archivos SQL", "*.sql")], title="Guardar respaldo de la base de datos")
        if ruta_respaldo_db:
            descargar_archivo("bd", ruta_respaldo_db)
    def descargar_imagenes_qr():
        ruta_zip = filedialog.asksaveasfilename(defaultextension=".zip", filetypes=[("Archivo ZIP", "*.zip")], title="Guardar imágenes de QR como ZIP")
        if not ruta_zip:
            return
        descargar_archivo("qr", ruta_zip)
    btn_descargar_excel = tk.Button(ventana_descargar, text="Descargar Excel", command=descargar_excel, width=20, height=2, bg="blue", fg="white", font=("Arial", 10))
    btn_descargar_excel.pack(pady=5)
    btn_descargar_bd = tk.Button(ventana_descargar, text="Descargar BD", command=descargar_bd, width=20, height=2, bg="darkblue", fg="white", font=("Arial", 10))
    btn_descargar_bd.pack(pady=5)
    btn_descargar_imagenes = tk.Button(ventana_descargar, text="Descargar Imágenes", command=descargar_imagenes_qr, width=20, height=2, bg="purple", fg="white", font=("Arial", 10))
    btn_descargar_imagenes.pack(pady=5)
    btn_cerrar = tk.Button(ventana_descargar, text="Cerrar", command=ventana_descargar.destroy, width=20, height=2, bg="red", fg="white", font=("Arial", 10))
    btn_cerrar.pack(pady=5)

# Ensure the "Descargar Todo" button is present
btn_descargar_todo = tk.Button(frame_botones, text="Descargar Datos", command=descargar_todo, width=20, height=2, bg="darkblue", fg="white", font=("Arial", 12, "bold"))

# Replace individual import buttons with a single "Importar Datos" button
def importar_datos():
    # Create a new window for import options
    ventana_importar = Toplevel(root)
    ventana_importar.title("Importar Datos")
    ventana_importar.geometry("400x300")

    label_titulo = tk.Label(ventana_importar, text="Selecciona qué deseas importar:", font=("Arial", 14))
    label_titulo.pack(pady=10)

    def importar_excel():
        try:
            ruta_excel = filedialog.askopenfilename(
                filetypes=[("Archivos de Excel", "*.xlsx")],
                title="Seleccionar archivo de Excel"
            )
            if not ruta_excel:
                return

            wb = load_workbook(ruta_excel)
            ws = wb.active

            columnas_esperadas = ["Nombre", "Apellido", "Número de Alumno"]
            columnas_excel = [cell.value for cell in ws[1] if cell.value]

            if not all(col in columnas_excel for col in columnas_esperadas):
                messagebox.showerror("Error", "El archivo Excel no tiene el formato adecuado.")
                return

            shutil.copy(ruta_excel, excel_file)
            messagebox.showinfo("Éxito", "El archivo Excel ha sido importado correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo importar el archivo Excel: {e}")

    def importar_base_datos():
        try:
            ruta_sql = filedialog.askopenfilename(
                filetypes=[("Archivos SQL", "*.sql")],
                title="Seleccionar archivo de base de datos SQL"
            )
            if not ruta_sql:
                return

            conexion = pymysql.connect(
                host="localhost",
                user="root",
                password="1234",
                database="asistencia"
            )
            cursor = conexion.cursor()
            with open(ruta_sql, "r") as archivo_sql:
                comandos = archivo_sql.read()
                for comando in comandos.split(";"):
                    if comando.strip():
                        cursor.execute(comando)
            conexion.commit()
            cursor.close()
            conexion.close()

            messagebox.showinfo("Éxito", "La base de datos ha sido importada correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo importar la base de datos: {e}")

    def importar_imagenes_qr():
        try:
            ruta_zip = filedialog.askopenfilename(
                filetypes=[("Archivo ZIP", "*.zip")],
                title="Seleccionar archivo ZIP con imágenes de QR"
            )
            if not ruta_zip:
                return

            with zipfile.ZipFile(ruta_zip, 'r') as zip_ref:
                zip_ref.extractall(qr_folder)

            messagebox.showinfo("Éxito", "Las imágenes de QR han sido importadas correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron importar las imágenes de QR: {e}")

    btn_importar_excel = tk.Button(ventana_importar, text="Importar Excel", command=importar_excel, width=20, height=2, bg="orange", fg="white", font=("Arial", 10))
    btn_importar_excel.pack(pady=5)

    btn_importar_db = tk.Button(ventana_importar, text="Importar Base de Datos", command=importar_base_datos, width=20, height=2, bg="purple", fg="white", font=("Arial", 10))
    btn_importar_db.pack(pady=5)

    btn_importar_qr = tk.Button(ventana_importar, text="Importar Imágenes de QR", command=importar_imagenes_qr, width=20, height=2, bg="blue", fg="white", font=("Arial", 10))
    btn_importar_qr.pack(pady=5)

    btn_cerrar = tk.Button(ventana_importar, text="Cerrar", command=ventana_importar.destroy, width=20, height=2, bg="red", fg="white", font=("Arial", 10))
    btn_cerrar.pack(pady=5)

# Replace the individual import buttons with a single "Importar Datos" button
btn_importar_datos = tk.Button(frame_botones, text="Importar Datos", command=importar_datos, width=20, height=2, bg="darkgreen", fg="white", font=("Arial", 12, "bold"))
btn_importar_datos.grid(row=2, column=0, padx=10, pady=20)  # Increased pady for better separation

# Ajustar la posición de los botones "Importar Datos" y "Descargar Datos" para que estén en la misma altura
btn_importar_datos.grid(row=2, column=0, padx=10, pady=20)  # Mover "Importar Datos" a la misma fila
btn_descargar_todo.grid(row=2, column=1, padx=10, pady=20)  # Alinear "Descargar Datos" en la misma fila

frame_generar_usuario = tk.Frame(root, bg="lightblue")
frame_generar_usuario.pack(pady=20)

label_nombre = tk.Label(frame_generar_usuario, text="Nombre(s):", bg="lightblue", fg="darkblue", font=("Arial", 12))
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(frame_generar_usuario, width=30)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

label_apellido = tk.Label(frame_generar_usuario, text="Apellido(s):", bg="lightblue", fg="darkblue", font=("Arial", 12))
label_apellido.grid(row=1, column=0, padx=10, pady=5)
entry_apellido = tk.Entry(frame_generar_usuario, width=30)
entry_apellido.grid(row=1, column=1, padx=10, pady=5)

entry_segundo_apellido = tk.Entry(frame_generar_usuario, width=30)
entry_segundo_apellido.grid(row=2, column=1, padx=10, pady=5)
label_segundo_apellido = tk.Label(frame_generar_usuario, text="Segundo Apellido:", bg="lightblue", fg="darkblue", font=("Arial", 12))
label_segundo_apellido.grid(row=2, column=0, padx=10, pady=5)

btn_generar_usuario = tk.Button(frame_generar_usuario, text="Generar Usuario", command=generar_usuario_desde_interfaz, bg="orange", fg="white", font=("Arial", 12, "bold"))
btn_generar_usuario.grid(row=3, column=0, columnspan=2, padx=10, pady=10)

# Asegurar que el botón "Mostrar Todos los QRs" esté visible y correctamente posicionado
btn_mostrar_qrs = tk.Button(root, text="Mostrar Todos los QRs", command=mostrar_todos_qrs, width=20, height=2, bg="purple", fg="white", font=("Arial", 12, "bold"))
btn_mostrar_qrs.pack(pady=20)  # Ajustar el espaciado para que sea consistente con el diseño general

# Add a footer
footer = tk.Label(root, text="Sistema de Asistencia con QR - 2025", bg="lightblue", fg="darkblue", font=("Arial", 10))
footer.pack(side=tk.BOTTOM, pady=10)

root.mainloop()