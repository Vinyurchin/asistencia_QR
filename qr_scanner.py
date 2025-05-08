import cv2
import pymysql
from pyzbar.pyzbar import decode
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import tkinter as tk
import json
from tkinter import messagebox, filedialog
import shutil
import os

excel_file = "asistencias.xlsx"

# Leer la configuración de la base de datos
with open("config.json", "r") as config_file:
    config = json.load(config_file)

conexion = pymysql.connect(
    host=config["host"],
    user=config["user"],
    password=config["password"],
    database=config["database"]
)

# Intentar cargar el archivo, si no existe, crearlo
try:
    wb = load_workbook(excel_file)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Apellido", "Número de Alumno"])  # Encabezados base
    wb.save(excel_file)
except PermissionError:
    print(f"Error: El archivo {excel_file} está abierto. Por favor, ciérralo e intenta de nuevo.")
    exit(1)

# Función para registrar asistencia correctamente en la columna de la fecha actual
def registrar_asistencia_excel(usuario_id, nombre, apellido):
    # Update the date format to include the month name and apply colors for better differentiation
    fecha_actual = datetime.now().strftime("%d-%B-%Y")  # Example: 07-Mayo-2025

    # Obtener la lista de encabezados
    encabezados = [cell.value for cell in ws[1] if cell.value]  # Ignorar celdas vacías

    # Verificar si la fecha ya existe en los encabezados
    if fecha_actual not in encabezados:
        col_index = len(encabezados) + 1  # Nueva columna para la fecha
        ws.cell(row=1, column=col_index, value=fecha_actual)
        wb.save(excel_file)
    else:
        col_index = encabezados.index(fecha_actual) + 1  # Buscar la columna correcta

    # Buscar la fila del usuario
    for row in ws.iter_rows(min_row=2, max_col=len(encabezados) + 1, values_only=False):
        if row[0].value == nombre and row[1].value == apellido and row[2].value == usuario_id:
            # Verificar si la celda correspondiente está marcada
            celda_asistencia = ws.cell(row=row[0].row, column=col_index)
            if celda_asistencia.value is not None:
                # Si la celda está marcada, confirmar que tiene el valor correcto
                if celda_asistencia.value != "Presente":
                    celda_asistencia.value = "Presente"
                    # Apply colors to differentiate dates
                    colores_mes = {
                        "Enero": "FFCCCC",
                        "Febrero": "FFCC99",
                        "Marzo": "FFFF99",
                        "Abril": "CCFFCC",
                        "Mayo": "99CCFF",
                        "Junio": "CCCCFF",
                        "Julio": "FF99CC",
                        "Agosto": "FF9966",
                        "Septiembre": "FFFF66",
                        "Octubre": "99FF99",
                        "Noviembre": "66CCFF",
                        "Diciembre": "CC99FF"
                    }

                    mes_actual = datetime.now().strftime("%B")
                    color = colores_mes.get(mes_actual, "FFFFFF")  # Default color: white

                    # When marking attendance, apply the color to the cell
                    celda_asistencia.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    wb.save(excel_file)
                    print(f"Se corrigió la asistencia en Excel para: {nombre} {apellido}")
                else:
                    print(f"Asistencia ya registrada en Excel para: {nombre} {apellido}")
                return

            # Si la celda está vacía, marcar asistencia
            celda_asistencia.value = "Presente"
            # Apply colors to differentiate dates
            colores_mes = {
                "Enero": "FFCCCC",
                "Febrero": "FFCC99",
                "Marzo": "FFFF99",
                "Abril": "CCFFCC",
                "Mayo": "99CCFF",
                "Junio": "CCCCFF",
                "Julio": "FF99CC",
                "Agosto": "FF9966",
                "Septiembre": "FFFF66",
                "Octubre": "99FF99",
                "Noviembre": "66CCFF",
                "Diciembre": "CC99FF"
            }

            mes_actual = datetime.now().strftime("%B")
            color = colores_mes.get(mes_actual, "FFFFFF")  # Default color: white

            # When marking attendance, apply the color to the cell
            celda_asistencia.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            wb.save(excel_file)  # Save the changes to the Excel file
            print(f"Asistencia guardada en Excel: {nombre} {apellido}")
            return

    # Si el usuario no está en la lista, agregarlo
    new_row = [nombre, apellido, usuario_id] + [None] * (col_index - 3)
    ws.append(new_row)
    ws.cell(row=ws.max_row, column=col_index, value="Presente")
    # Apply colors to differentiate dates
    colores_mes = {
        "Enero": "FFCCCC",
        "Febrero": "FFCC99",
        "Marzo": "FFFF99",
        "Abril": "CCFFCC",
        "Mayo": "99CCFF",
        "Junio": "CCCCFF",
        "Julio": "FF99CC",
        "Agosto": "FF9966",
        "Septiembre": "FFFF66",
        "Octubre": "99FF99",
        "Noviembre": "66CCFF",
        "Diciembre": "CC99FF"
    }

    mes_actual = datetime.now().strftime("%B")
    color = colores_mes.get(mes_actual, "FFFFFF")  # Default color: white

    # When marking attendance, apply the color to the cell
    ws.cell(row=ws.max_row, column=col_index).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    wb.save(excel_file)  # Save the changes to the Excel file
    print(f"Asistencia guardada en Excel: {nombre} {apellido}")

# Función para descargar el archivo Excel
def descargar_excel():
    if not os.path.exists(excel_file):
        messagebox.showerror("Error", "El archivo de Excel no existe.")
        return

    # Guardar el archivo actualizado antes de descargarlo
    try:
        wb.save(excel_file)  # Asegúrate de guardar los cambios en el archivo
    except PermissionError:
        messagebox.showerror("Error", f"El archivo {excel_file} está abierto. Por favor, ciérralo e intenta de nuevo.")
        return

    ruta_destino = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos de Excel", "*.xlsx")],
        title="Guardar archivo de Excel"
    )
    if ruta_destino:
        try:
            shutil.copy(excel_file, ruta_destino)  # Copiar el archivo en lugar de moverlo
            messagebox.showinfo("Éxito", f"Archivo guardado en: {ruta_destino}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el archivo: {e}")

# Función para descargar el archivo Excel desde otro módulo
def descargar_excel_desde_qr_scanner():
    if not os.path.exists(excel_file):
        raise FileNotFoundError("El archivo de Excel no existe.")

    # Guardar el archivo actualizado antes de descargarlo
    try:
        wb.save(excel_file)  # Asegúrate de guardar los cambios en el archivo
    except PermissionError:
        raise PermissionError(f"El archivo {excel_file} está abierto. Por favor, ciérralo e intenta de nuevo.")

# Función principal para iniciar el escaneo de QR
def iniciar_escaneo_qr():
    try:
        # Mostrar una alerta al usuario para indicar cómo cerrar la ventana
        root = tk.Tk()
        root.withdraw()  # Ocultar la ventana principal de tkinter
        messagebox.showinfo("Instrucciones", "Presiona 'q' para cerrar la ventana de la cámara.")

        # Conectar a MySQL
        conexion = pymysql.connect(
            host="localhost",
            user="root",
            password="1234",
            database="asistencia"
        )
        cursor = conexion.cursor()

        # Inicializa la cámara
        cap = cv2.VideoCapture(0)
        procesados = set()  # Lista para evitar múltiples registros del mismo QR

        # Variable de control para detener el bucle
        escaneando = True
    
        while escaneando:
            ret, frame = cap.read()
            if not ret:
                break

            # Decodifica los QR en la imagen
            for qr in decode(frame):
                try:
                    if not qr.data or not hasattr(qr, 'rect'):
                        continue

                    qr_data = qr.data.decode("utf-8")
                    qr_rect = qr.rect

                    # Evitar procesar el mismo QR repetidamente
                    if qr_data in procesados:
                        continue

                    # Dibuja un rectángulo alrededor del QR
                    cv2.rectangle(frame, (qr_rect.left, qr_rect.top),
                                  (qr_rect.left + qr_rect.width, qr_rect.top + qr_rect.height),
                                  (0, 255, 0), 3)

                    # Buscar el código QR en la base de datos
                    cursor.execute("SELECT id, nombre, apellido FROM usuarios WHERE qr_code = %s", (qr_data,))
                    usuario = cursor.fetchone()

                    if usuario:
                        usuario_id, nombre, apellido = usuario

                        # Validar datos del usuario
                        if not usuario_id or not nombre or not apellido:
                            mensaje = "Datos del usuario inválidos."
                            cv2.putText(frame, mensaje, (qr_rect.left, qr_rect.top - 10),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                            continue

                        fecha_actual = datetime.now().strftime('%Y-%m-%d')

                        # Verificar si ya tiene asistencia hoy
                        cursor.execute("SELECT * FROM asistencia WHERE usuario_id = %s AND DATE(fecha_asistencia) = %s", (usuario_id, fecha_actual))
                        asistencia_existente = cursor.fetchone()

                        if asistencia_existente:
                            mensaje = f"{nombre} {apellido} ya tiene asistencia registrada hoy."
                            print(mensaje)
                            cv2.putText(frame, mensaje, (qr_rect.left, qr_rect.top - 10),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                        else:
                            # Registrar asistencia en MySQL
                            cursor.execute("INSERT INTO asistencia (usuario_id, fecha_asistencia) VALUES (%s, NOW())", (usuario_id,))
                            conexion.commit()

                            # Guardar en Excel
                            registrar_asistencia_excel(usuario_id, nombre, apellido)

                            mensaje = f"Asistencia registrada: {nombre} {apellido}"
                            print(mensaje)
                            cv2.putText(frame, mensaje, (qr_rect.left, qr_rect.top - 10),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

                            # Mostrar una confirmación emergente que se cierra automáticamente
                            popup = tk.Toplevel()
                            popup.title("Asistencia Registrada")
                            tk.Label(popup, text=mensaje, font=("Arial", 14), fg="green").pack(pady=10, padx=10)
                            popup.after(3000, popup.destroy)  # Cerrar automáticamente después de 3 segundos

                        # Agregar el QR procesado a la lista
                        procesados.add(qr_data)
                    else:
                        mensaje = "Código QR no registrado"
                        print(mensaje)
                        cv2.putText(frame, mensaje, (qr_rect.left, qr_rect.top - 10),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)

                except pymysql.MySQLError as e:
                    mensaje = f"Error en la base de datos: {e}"
                    print(mensaje)
                    cv2.putText(frame, mensaje, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

            # Muestra la cámara en tiempo real
            cv2.imshow("Escáner QR", frame)

            # Detectar si la ventana fue cerrada
            if cv2.getWindowProperty("Escáner QR", cv2.WND_PROP_VISIBLE) < 1:
                escaneando = False

            # Presiona 'q' para salir manualmente
            if cv2.waitKey(1) & 0xFF == ord("q"):
                escaneando = False

        # Libera la cámara y cierra la ventana
        cap.release()
        cv2.destroyAllWindows()

        # Cierra la conexión a la base de datos
        cursor.close()
        conexion.close()

    except Exception as e:
        print(f"Error al iniciar el escaneo de QR: {e}")

# Asegúrate de que este código solo se ejecute si el archivo se ejecuta directamente
if __name__ == "__main__":
    iniciar_escaneo_qr()